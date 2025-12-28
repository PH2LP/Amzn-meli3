#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_excel_v2.py - VERSIÓN MEJORADA
Genera Excel profesional con 3 hojas optimizadas:
1. Dashboard - Vista ejecutiva con KPIs y gráficos
2. Ventas - Detalle completo de transacciones
3. Por Día - Resumen diario agregado

Características:
- Diseño limpio y profesional
- Gráficos dimensionados correctamente
- Optimizado para ejecución cada 1 minuto
- Auto-upload a Dropbox
"""

import sqlite3
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import os

# ═══════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════

DB_PATH = "storage/sales_tracking.db"
EXCEL_FILENAME = f"{datetime.now().strftime('%Y%m%d')}_VENTAS_MERCADOLIBRE.xlsx"

# Detectar entorno (servidor o local)
if os.path.exists("/opt/amz-ml-system"):
    EXCEL_PATH = f"/opt/amz-ml-system/storage/{EXCEL_FILENAME}"
else:
    EXCEL_PATH = f"/Users/felipemelucci/Desktop/{EXCEL_FILENAME}"

# Colores corporativos
COLOR_PRIMARY = "1F4E78"      # Azul principal
COLOR_SUCCESS = "006100"      # Verde (ganancias)
COLOR_DANGER = "CC0000"       # Rojo (costos)
COLOR_INFO = "0066CC"         # Azul info
COLOR_HEADER = "4472C4"       # Azul headers
COLOR_LIGHT_GRAY = "F2F2F2"   # Gris claro (alternancia)

# ═══════════════════════════════════════════════════════════════
# UTILIDADES
# ═══════════════════════════════════════════════════════════════

def get_sales_data():
    """Obtiene datos de ventas desde la DB"""
    conn = sqlite3.connect(DB_PATH)

    query = """
        SELECT
            sale_date as 'Fecha',
            marketplace as 'MKT',
            product_title as 'Producto',
            quantity as 'Cant',
            sale_price_usd as 'Precio Venta',
            ml_fee as 'Fee ML',
            shipping_cost as 'Envío',
            net_proceeds as 'Neto ML',
            amazon_cost as 'Costo AMZ',
            fulfillment_fee as '3PL',
            total_cost as 'Total Costo',
            profit as 'GANANCIA',
            profit_margin as 'Margen %',
            asin as 'ASIN',
            ml_item_id as 'ML ID'
        FROM sales
        ORDER BY sale_date DESC
    """

    df = pd.read_sql_query(query, conn)
    conn.close()

    # Limpiar y formatear datos
    if len(df) > 0:
        df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce').dt.strftime('%Y-%m-%d %H:%M')

        # Convertir NaN a 0 en columnas numéricas
        numeric_cols = ['Cant', 'Precio Venta', 'Fee ML', 'Envío', 'Neto ML',
                       'Costo AMZ', '3PL', 'Total Costo', 'GANANCIA', 'Margen %']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    return df

def calculate_kpis(df):
    """Calcula KPIs principales"""
    if len(df) == 0:
        return {
            'total_ventas': 0,
            'total_revenue': 0,
            'total_profit': 0,
            'total_costs': 0,
            'total_ml_fees': 0,
            'total_amazon_cost': 0,
            'total_3pl': 0,
            'roi': 0,
            'avg_ticket': 0,
            'avg_margin': 0,
        }

    total_revenue = df['Precio Venta'].sum()
    total_costs = df['Total Costo'].sum()

    return {
        'total_ventas': len(df),
        'total_revenue': total_revenue,
        'total_profit': df['GANANCIA'].sum(),
        'total_costs': total_costs,
        'total_ml_fees': df['Fee ML'].sum(),
        'total_amazon_cost': df['Costo AMZ'].sum(),
        'total_3pl': df['3PL'].sum(),
        'roi': (df['GANANCIA'].sum() / total_costs * 100) if total_costs > 0 else 0,
        'avg_ticket': total_revenue / len(df) if len(df) > 0 else 0,
        'avg_margin': df['Margen %'].mean(),
    }

def apply_table_style(ws, start_row, end_row, start_col, end_col):
    """Aplica estilo profesional a una tabla"""
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # Aplicar bordes y alternancia de colores
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

            # Alternancia de colores (solo filas de datos, no headers)
            if row_idx > start_row and (row_idx - start_row) % 2 == 0:
                cell.fill = PatternFill(start_color=COLOR_LIGHT_GRAY,
                                       end_color=COLOR_LIGHT_GRAY,
                                       fill_type="solid")

# ═══════════════════════════════════════════════════════════════
# HOJA 1: DASHBOARD
# ═══════════════════════════════════════════════════════════════

def create_dashboard(wb, df, kpis):
    """Crea la hoja Dashboard con KPIs y gráficos"""
    ws = wb.create_sheet("Dashboard", 0)

    # ═══ HEADER PRINCIPAL ═══
    ws['A1'] = '📊 VENTAS AMAZON → MERCADOLIBRE - DASHBOARD EJECUTIVO'
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
    ws.merge_cells('A1:J1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # ═══ ÚLTIMA ACTUALIZACIÓN ═══
    ws['A2'] = f'🕐 Última actualización: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(name='Arial', size=10, italic=True, color="666666")
    ws.merge_cells('A2:J2')
    ws['A2'].alignment = Alignment(horizontal='center')

    # ═══ SECCIÓN KPIs PRINCIPALES ═══
    ws['A4'] = '💰 RESUMEN FINANCIERO'
    ws['A4'].font = Font(name='Arial', size=14, bold=True, color=COLOR_PRIMARY)
    ws.merge_cells('A4:D4')

    kpi_data = [
        ['Total Ventas:', kpis['total_ventas'], 'unidades', None],
        ['Revenue Total:', kpis['total_revenue'], 'USD', COLOR_SUCCESS],
        ['Ganancia Total:', kpis['total_profit'], 'USD', COLOR_SUCCESS],
        ['ROI:', kpis['roi'], '%', COLOR_INFO],
        ['Ticket Promedio:', kpis['avg_ticket'], 'USD', COLOR_INFO],
        ['Margen Promedio:', kpis['avg_margin'], '%', COLOR_INFO],
    ]

    row = 5
    for label, value, unit, color in kpi_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True)

        if unit == 'USD':
            ws[f'B{row}'] = f'${value:,.2f}'
        elif unit == '%':
            ws[f'B{row}'] = f'{value:.2f}%'
        else:
            ws[f'B{row}'] = int(value)

        ws[f'B{row}'].font = Font(name='Arial', size=12, bold=True,
                                   color=color if color else "000000")
        row += 1

    # ═══ DESGLOSE DE COSTOS ═══
    ws['E4'] = '📉 DESGLOSE DE COSTOS'
    ws['E4'].font = Font(name='Arial', size=14, bold=True, color=COLOR_PRIMARY)
    ws.merge_cells('E4:G4')

    costs_data = [
        ['Comisiones ML:', kpis['total_ml_fees']],
        ['Costos Amazon:', kpis['total_amazon_cost']],
        ['3PL Fulfillment:', kpis['total_3pl']],
        ['Total Costos:', kpis['total_costs']],
    ]

    row = 5
    for label, value in costs_data:
        ws[f'E{row}'] = label
        ws[f'E{row}'].font = Font(name='Arial', size=11, bold=True)
        ws[f'F{row}'] = f'${value:,.2f}'
        ws[f'F{row}'].font = Font(name='Arial', size=11,
                                   color=COLOR_DANGER if 'Total' in label else "000000")
        row += 1

    # ═══ GRÁFICO 1: VENTAS POR DÍA (Line Chart) ═══
    if len(df) > 0:
        # Preparar datos diarios
        df_copy = df.copy()
        df_copy['Fecha_Solo'] = pd.to_datetime(df_copy['Fecha'], errors='coerce').dt.date
        daily_sales = df_copy.groupby('Fecha_Solo').agg({
            'Producto': 'count',
            'GANANCIA': 'sum'
        }).reset_index().sort_values('Fecha_Solo')

        # Limitar a últimos 15 días
        daily_sales = daily_sales.tail(15)

        ws['H4'] = '📈 VENTAS DIARIAS (ÚLTIMOS 15 DÍAS)'
        ws['H4'].font = Font(name='Arial', size=12, bold=True, color=COLOR_PRIMARY)
        ws.merge_cells('H4:J4')

        # Escribir datos para el gráfico
        ws['H5'] = 'Fecha'
        ws['I5'] = 'Ventas'
        ws['J5'] = 'Ganancia'

        chart_row = 6
        for _, row_data in daily_sales.iterrows():
            ws[f'H{chart_row}'] = str(row_data['Fecha_Solo'])
            ws[f'I{chart_row}'] = row_data['Producto']
            ws[f'J{chart_row}'] = row_data['GANANCIA']
            chart_row += 1

        # Crear Line Chart
        chart = LineChart()
        chart.title = "Ventas y Ganancia por Día"
        chart.style = 13
        chart.y_axis.title = 'Cantidad / USD'
        chart.x_axis.title = 'Fecha'

        data = Reference(ws, min_col=9, min_row=5, max_row=chart_row-1, max_col=10)
        cats = Reference(ws, min_col=8, min_row=6, max_row=chart_row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 12
        chart.width = 20

        ws.add_chart(chart, "A12")

    # ═══ GRÁFICO 2: TOP 5 PRODUCTOS (Bar Chart) ═══
    if len(df) > 0:
        top_products = df.nlargest(5, 'GANANCIA')[['Producto', 'GANANCIA']]

        ws['H22'] = '🏆 TOP 5 PRODUCTOS MÁS RENTABLES'
        ws['H22'].font = Font(name='Arial', size=12, bold=True, color=COLOR_PRIMARY)
        ws.merge_cells('H22:J22')

        ws['H23'] = 'Producto'
        ws['I23'] = 'Ganancia'

        chart_row = 24
        for _, prod in top_products.iterrows():
            ws[f'H{chart_row}'] = prod['Producto'][:30]
            ws[f'I{chart_row}'] = prod['GANANCIA']
            chart_row += 1

        # Crear Bar Chart horizontal
        chart2 = BarChart()
        chart2.type = "bar"  # horizontal
        chart2.title = "Top 5 Productos por Ganancia"
        chart2.style = 10

        data2 = Reference(ws, min_col=9, min_row=23, max_row=chart_row-1)
        cats2 = Reference(ws, min_col=8, min_row=24, max_row=chart_row-1)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.height = 10
        chart2.width = 15

        ws.add_chart(chart2, "A29")

    # ═══ GRÁFICO 3: PIE CHART - REVENUE VS COSTOS ═══
    ws['E12'] = '💵 DISTRIBUCIÓN DE REVENUE'
    ws['E12'].font = Font(name='Arial', size=12, bold=True, color=COLOR_PRIMARY)
    ws.merge_cells('E12:G12')

    ws['E13'] = 'Categoría'
    ws['F13'] = 'Monto'
    ws['E14'] = 'Ganancia Neta'
    ws['F14'] = kpis['total_profit']
    ws['E15'] = 'Costos Totales'
    ws['F15'] = kpis['total_costs']

    pie = PieChart()
    labels = Reference(ws, min_col=5, min_row=14, max_row=15)
    data = Reference(ws, min_col=6, min_row=13, max_row=15)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Revenue vs Costos"
    pie.height = 12
    pie.width = 15

    ws.add_chart(pie, "E17")

    # ═══ AJUSTAR ANCHOS DE COLUMNA ═══
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12

# ═══════════════════════════════════════════════════════════════
# HOJA 2: VENTAS (Detalle completo)
# ═══════════════════════════════════════════════════════════════

def create_ventas_sheet(wb, df):
    """Crea la hoja Ventas con detalle completo de transacciones"""
    ws = wb.create_sheet("Ventas", 1)

    # Escribir dataframe completo
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

    # Headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Aplicar estilo de tabla
    if len(df) > 0:
        apply_table_style(ws, 1, len(df) + 1, 1, len(headers))

    # Formato de moneda y números
    for row_idx in range(2, len(df) + 2):
        # Precio Venta, Fee ML, Envío, Neto ML, Costo AMZ, 3PL, Total Costo, GANANCIA
        for col in [5, 6, 7, 8, 9, 10, 11, 12]:
            ws.cell(row=row_idx, column=col).number_format = '$#,##0.00'
        # Margen %
        ws.cell(row=row_idx, column=13).number_format = '0.00"%"'

    # Ajustar anchos de columna
    column_widths = {
        'A': 18, 'B': 8, 'C': 50, 'D': 6, 'E': 13,
        'F': 11, 'G': 10, 'H': 12, 'I': 12, 'J': 10,
        'K': 12, 'L': 13, 'M': 11, 'N': 15, 'O': 15
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Habilitar filtros automáticos
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# ═══════════════════════════════════════════════════════════════
# HOJA 3: POR DÍA (Resumen diario)
# ═══════════════════════════════════════════════════════════════

def create_daily_sheet(wb, df):
    """Crea la hoja Por Día con resumen diario agregado"""
    ws = wb.create_sheet("Por Día", 2)

    if len(df) == 0:
        ws['A1'] = 'No hay datos disponibles'
        return

    # Preparar datos diarios
    df_copy = df.copy()
    df_copy['Fecha_Solo'] = pd.to_datetime(df_copy['Fecha'], errors='coerce').dt.date

    daily_summary = df_copy.groupby('Fecha_Solo').agg({
        'Producto': 'count',           # Ventas (cantidad)
        'Precio Venta': 'sum',         # Revenue
        'GANANCIA': 'sum',             # Ganancia
        'Margen %': 'mean',            # Margen promedio
        'Costo AMZ': 'sum',            # Costo Amazon total
        'Fee ML': 'sum',               # Fee ML total
        'Envío': 'sum',                # Envío total
        'ASIN': 'nunique'              # Productos únicos
    }).reset_index()

    # Renombrar columnas
    daily_summary.columns = ['Fecha', 'Ventas', 'Revenue', 'Ganancia',
                            'Margen %', 'Costo AMZ', 'Fee ML', 'Envío', 'Productos Únicos']

    # Ordenar descendente (más reciente primero)
    daily_summary = daily_summary.sort_values('Fecha', ascending=False)

    # Headers
    headers = list(daily_summary.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Escribir datos
    for r_idx, row in enumerate(daily_summary.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

    # Aplicar estilo de tabla
    apply_table_style(ws, 1, len(daily_summary) + 1, 1, len(headers))

    # Formato de números
    for row_idx in range(2, len(daily_summary) + 2):
        # Revenue, Ganancia, Costo AMZ, Fee ML, Envío
        for col in [3, 4, 6, 7, 8]:
            ws.cell(row=row_idx, column=col).number_format = '$#,##0.00'
        # Margen %
        ws.cell(row=row_idx, column=5).number_format = '0.00"%"'

    # Fila de totales
    total_row = len(daily_summary) + 2
    ws[f'A{total_row}'] = 'TOTAL'
    ws[f'A{total_row}'].font = Font(name='Arial', size=11, bold=True)
    ws[f'B{total_row}'] = daily_summary['Ventas'].sum()
    ws[f'C{total_row}'] = daily_summary['Revenue'].sum()
    ws[f'D{total_row}'] = daily_summary['Ganancia'].sum()
    ws[f'E{total_row}'] = daily_summary['Margen %'].mean()
    ws[f'F{total_row}'] = daily_summary['Costo AMZ'].sum()
    ws[f'G{total_row}'] = daily_summary['Fee ML'].sum()
    ws[f'H{total_row}'] = daily_summary['Envío'].sum()

    # Formato totales
    for col in [3, 4, 6, 7, 8]:
        ws.cell(row=total_row, column=col).number_format = '$#,##0.00'
        ws.cell(row=total_row, column=col).font = Font(name='Arial', size=11, bold=True, color=COLOR_SUCCESS)
    ws.cell(row=total_row, column=5).number_format = '0.00"%"'
    ws.cell(row=total_row, column=5).font = Font(name='Arial', size=11, bold=True, color=COLOR_INFO)

    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 16

# ═══════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ═══════════════════════════════════════════════════════════════

def generate_excel_v2():
    """Genera Excel mejorado con 3 hojas"""
    print("🔄 Generando Excel V2...")

    # Obtener datos
    df = get_sales_data()
    print(f"   📊 {len(df)} ventas cargadas desde DB")

    # Calcular KPIs
    kpis = calculate_kpis(df)

    # Crear workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remover hoja default

    # Crear las 3 hojas
    print("   📄 Creando hoja Dashboard...")
    create_dashboard(wb, df, kpis)

    print("   📄 Creando hoja Ventas...")
    create_ventas_sheet(wb, df)

    print("   📄 Creando hoja Por Día...")
    create_daily_sheet(wb, df)

    # Guardar Excel
    wb.save(EXCEL_PATH)
    print(f"✅ Excel V2 creado: {EXCEL_PATH}")
    print(f"   Total ventas: {kpis['total_ventas']}")
    if kpis['total_ventas'] > 0:
        print(f"   Ganancia total: ${kpis['total_profit']:,.2f}")
        print(f"   ROI: {kpis['roi']:.2f}%")

    return EXCEL_PATH

def upload_to_dropbox(excel_path):
    """Sube el Excel a Dropbox con auto-refresh de token"""
    try:
        import sys
        from pathlib import Path

        # Add project root to path
        script_dir = Path(__file__).parent
        sys.path.insert(0, str(script_dir))

        # Import dropbox auth utility
        from dropbox_auth import get_dropbox_client
        import dropbox

        # Get Dropbox client
        dbx = get_dropbox_client()
        if not dbx:
            print("⚠️  No se pudo obtener cliente Dropbox - Excel no se subió")
            return

        # Upload Excel
        with open(excel_path, 'rb') as f:
            dropbox_path = f"/VENTAS_MERCADOLIBRE.xlsx"
            dbx.files_upload(f.read(), dropbox_path, mode=dropbox.files.WriteMode.overwrite)
            file_size = os.path.getsize(excel_path) / 1024  # KB
            print(f"✅ Excel subido a Dropbox: {dropbox_path} ({file_size:.1f} KB)")

        # Upload DBs también
        db_files = [
            ("storage/sales_tracking.db", "/sales_tracking.db"),
            ("storage/listings_database.db", "/listings_database.db")
        ]

        for local_path, dropbox_path in db_files:
            if os.path.exists(local_path):
                with open(local_path, 'rb') as f:
                    dbx.files_upload(f.read(), dropbox_path, mode=dropbox.files.WriteMode.overwrite)
                    file_size = os.path.getsize(local_path) / 1024
                    print(f"✅ DB subida: {dropbox_path} ({file_size:.1f} KB)")

    except Exception as e:
        print(f"❌ Error subiendo a Dropbox: {str(e)}")

# ═══════════════════════════════════════════════════════════════
# EJECUCIÓN
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    try:
        excel_path = generate_excel_v2()
        upload_to_dropbox(excel_path)
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
