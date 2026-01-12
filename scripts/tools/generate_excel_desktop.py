#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_excel_desktop.py
Genera Excel profesional con ventas en el Desktop
"""

import sqlite3
import pandas as pd
import datetime
from datetime import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuración
DB_PATH = "storage/sales_tracking.db"
EXCEL_PATH = f"{dt.now().strftime('%Y%m%d')}_VENTAS_ML_V2.xlsx"

# Detectar si estamos en servidor o local
import os
if os.path.exists("/opt/amz-ml-system"):
    # Estamos en el servidor
    DESKTOP_PATH = f"/opt/amz-ml-system/storage/{EXCEL_PATH}"
else:
    # Estamos en local (Mac)
    DESKTOP_PATH = f"/Users/felipemelucci/Desktop/{EXCEL_PATH}"

def get_sales_count():
    """Obtiene el número total de ventas en la DB"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM sales")
        count = cursor.fetchone()[0]
        conn.close()
        return count
    except:
        return 0

def has_sales_changed():
    """Verifica si hay cambios en las ventas desde la última subida"""
    LAST_COUNT_FILE = "storage/.last_sales_count"

    current_count = get_sales_count()

    # Leer último count guardado
    try:
        with open(LAST_COUNT_FILE, 'r') as f:
            last_count = int(f.read().strip())
    except:
        last_count = -1  # Forzar subida en primera ejecución

    # Si hay cambios, guardar nuevo count
    if current_count != last_count:
        with open(LAST_COUNT_FILE, 'w') as f:
            f.write(str(current_count))
        return True

    return False

def create_professional_excel():
    """Crea Excel profesional en el Desktop"""

    # Conectar a DB
    conn = sqlite3.connect(DB_PATH)

    # Leer datos con TODAS las columnas (orden optimizado)
    df = pd.read_sql_query("""
        SELECT
            sale_date as 'Fecha',
            marketplace as 'MKT',
            product_title as 'Producto',
            quantity as 'Cant',
            sale_price_usd as 'Precio Venta',
            ml_fee as 'Fee ML',
            shipping_cost as 'Envío',
            net_proceeds as 'Neto ML',
            amazon_cost as 'Costo AMZ',
            fulfillment_fee as '3PL',
            total_cost as 'Total Costo',
            profit as 'GANANCIA',
            profit_margin as 'Margen %',
            asin as 'ASIN',
            order_id as 'Orden ML',
            ml_item_id as 'CBT ID',
            buyer_nickname as 'Comprador',
            country as 'País',
            status as 'Estado'
        FROM sales
        ORDER BY sale_date DESC
    """, conn)

    conn.close()

    # Si no hay datos, crear archivo vacío con headers
    if len(df) == 0:
        print("⚠️  No hay ventas aún. Creando Excel vacío...")

        # Crear DataFrame vacío con columnas
        df = pd.DataFrame(columns=[
            'Fecha', 'MKT', 'Producto', 'Cant',
            'Precio Venta', 'Fee ML', 'Envío', 'Neto ML',
            'Costo AMZ', '3PL', 'Total Costo',
            'GANANCIA', 'Margen %', 'ASIN', 'Orden ML', 'CBT ID',
            'Comprador', 'País', 'Estado'
        ])

    # Formatear fecha (maneja múltiples formatos: ISO y simples)
    if 'Fecha' in df.columns and len(df) > 0:
        df['Fecha'] = pd.to_datetime(df['Fecha'], format='mixed', errors='coerce')
        df['Fecha'] = df['Fecha'].apply(lambda x: x.strftime('%Y-%m-%d %H:%M') if pd.notna(x) else '')

    # Exportar a Excel
    df.to_excel(DESKTOP_PATH, index=False, sheet_name='Ventas')

    # Aplicar formato profesional
    wb = load_workbook(DESKTOP_PATH)
    ws = wb['Ventas']

    # Colores
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ganancia_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    costo_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Fonts
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    normal_font = Font(name='Arial', size=10)

    # Border
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # Formatear headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Ajustar anchos de columna (más espaciosos para evitar texto apretado)
    column_widths = {
        'A': 18,  # Fecha
        'B': 8,   # MKT
        'C': 50,  # Producto (más ancho)
        'D': 6,   # Cant
        'E': 13,  # Precio Venta
        'F': 11,  # Fee ML
        'G': 10,  # Envío
        'H': 12,  # Neto ML
        'I': 12,  # Costo AMZ
        'J': 8,   # 3PL
        'K': 13,  # Total Costo
        'L': 13,  # GANANCIA
        'M': 11,  # Margen %
        'N': 14,  # ASIN
        'O': 18,  # Orden ML
        'P': 16,  # CBT ID
        'Q': 22,  # Comprador
        'R': 12,  # País
        'S': 10,  # Estado
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Formatear celdas de datos con todas las líneas gruesas
    thick_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            # Todas las letras en negrita
            cell.font = Font(name='Arial', size=10, bold=True)

            # Todas las líneas gruesas
            cell.border = thick_border

            cell.alignment = Alignment(vertical='center')

            # Formatear números como moneda (columnas financieras)
            col_letter = get_column_letter(cell.column)
            # E=Precio Venta, F=Fee ML, G=Envío, H=Neto ML,
            # I=Costo AMZ, J=3PL, K=Total Costo, L=GANANCIA
            if col_letter in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                cell.number_format = '$#,##0.00'

            # Formatear porcentaje (M=Margen %)
            elif col_letter == 'M':
                cell.number_format = '0.00"%"'

            # Resaltar ganancia (L=GANANCIA)
            if col_letter == 'L':
                cell.fill = ganancia_fill
                cell.font = Font(name='Arial', size=10, bold=True, color="006100")

            # Resaltar costos (I=Costo AMZ, J=3PL, K=Total Costo)
            elif col_letter in ['I', 'J', 'K']:
                cell.fill = costo_fill

    # Fijar primera fila
    ws.freeze_panes = 'A2'

    # Agregar filtros
    ws.auto_filter.ref = ws.dimensions

    # Agregar hoja de resumen PRO con gráficos
    ws_summary = wb.create_sheet("Dashboard", 0)

    if len(df) > 0:
        from openpyxl.chart import BarChart, PieChart, Reference, LineChart
        from openpyxl.chart.label import DataLabelList

        # Filtrar solo ventas confirmadas para métricas (excluir canceladas)
        df_confirmed = df[df['Estado'] != 'cancelled'].copy()
        df_cancelled = df[df['Estado'] == 'cancelled'].copy()

        # Calcular métricas clave (solo ventas confirmadas)
        total_ventas = len(df_confirmed)
        total_cancelled = len(df_cancelled)
        total_revenue = df_confirmed['Precio Venta'].sum()
        total_ml_fees = df_confirmed['Fee ML'].sum()
        total_net_ml = df_confirmed['Neto ML'].sum()
        total_amazon_cost = df_confirmed['Costo AMZ'].sum()
        total_3pl = df_confirmed['3PL'].sum()
        total_costs = df_confirmed['Total Costo'].sum()
        total_profit = df_confirmed['GANANCIA'].sum()
        avg_margin = df_confirmed['Margen %'].mean()
        avg_ticket = df_confirmed['Precio Venta'].mean()
        roi = (total_profit / total_costs * 100) if total_costs > 0 else 0

        # Ventas por marketplace (solo confirmadas)
        by_marketplace = df_confirmed.groupby('MKT').agg({
            'GANANCIA': 'sum',
            'Producto': 'count'
        }).reset_index()

        # Top productos (solo confirmadas)
        top_products = df_confirmed.nlargest(5, 'GANANCIA')[['Producto', 'GANANCIA', 'Margen %']]

        # ═══ SECCIÓN 1: HEADER ═══
        ws_summary['A1'] = '📊 DASHBOARD DE VENTAS - AMAZON → MERCADOLIBRE'
        ws_summary['A1'].font = Font(name='Arial', size=18, bold=True, color="FFFFFF")
        ws_summary['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws_summary.merge_cells('A1:F1')
        ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_summary.row_dimensions[1].height = 30

        # ═══ SECCIÓN 2: KPIs PRINCIPALES ═══
        ws_summary['A3'] = '💰 RESUMEN FINANCIERO'
        ws_summary['A3'].font = Font(name='Arial', size=14, bold=True, color="1F4E78")

        kpis = [
            ['Total Ventas:', total_ventas, 'unidades'],
            ['Ventas Canceladas:', total_cancelled, 'unidades'],
            ['Revenue Total:', total_revenue, 'USD'],
            ['Ganancia Total:', total_profit, 'USD'],
            ['ROI:', roi, '%'],
            ['Ticket Promedio:', avg_ticket, 'USD'],
            ['Margen Promedio:', avg_margin, '%'],
        ]

        row = 4
        for label, value, unit in kpis:
            ws_summary[f'A{row}'] = label
            ws_summary[f'A{row}'].font = Font(name='Arial', size=11, bold=True)

            if unit == 'USD':
                ws_summary[f'B{row}'] = f'${value:,.2f}'
                ws_summary[f'B{row}'].font = Font(name='Arial', size=12, bold=True, color="006100")
            elif unit == '%':
                ws_summary[f'B{row}'] = f'{value:.2f}%'
                ws_summary[f'B{row}'].font = Font(name='Arial', size=12, bold=True, color="0066CC")
            else:
                ws_summary[f'B{row}'] = int(value)
                ws_summary[f'B{row}'].font = Font(name='Arial', size=12, bold=True)

            row += 1

        # ═══ SECCIÓN 3: DESGLOSE DE COSTOS ═══
        ws_summary['D3'] = '📉 DESGLOSE DE COSTOS'
        ws_summary['D3'].font = Font(name='Arial', size=14, bold=True, color="1F4E78")

        costs_breakdown = [
            ['Comisiones ML:', total_ml_fees],
            ['Costos Amazon:', total_amazon_cost],
            ['3PL Fulfillment:', total_3pl],
            ['Total Costos:', total_costs],
        ]

        row = 4
        for label, value in costs_breakdown:
            ws_summary[f'D{row}'] = label
            ws_summary[f'D{row}'].font = Font(name='Arial', size=11, bold=True)
            ws_summary[f'E{row}'] = f'${value:,.2f}'
            ws_summary[f'E{row}'].font = Font(name='Arial', size=11)
            if 'Total' in label:
                ws_summary[f'E{row}'].font = Font(name='Arial', size=12, bold=True, color="CC0000")
            row += 1

        # ═══ SECCIÓN 4: GRÁFICO DE FACTURACIÓN POR DÍA/MES/AÑO ═══
        ws_summary['A12'] = '📈 FACTURACIÓN EN EL TIEMPO'
        ws_summary['A12'].font = Font(name='Arial', size=14, bold=True, color="1F4E78")

        # Preparar datos temporales (solo confirmadas)
        df_time = df_confirmed.copy()
        df_time['Fecha_parsed'] = pd.to_datetime(df_time['Fecha'], format='mixed', errors='coerce')
        df_time = df_time[df_time['Fecha_parsed'].notna()].copy()
        df_time['Fecha_Solo'] = df_time['Fecha_parsed'].dt.date

        # Agrupar por día para el gráfico
        daily_revenue = df_time.groupby('Fecha_Solo').agg({
            'Precio Venta': 'sum'
        }).reset_index().sort_values('Fecha_Solo')

        # Escribir datos para el gráfico (últimos 30 días o todos si son menos)
        ws_summary['A14'] = 'Fecha'
        ws_summary['B14'] = 'Revenue'
        ws_summary['A14'].font = Font(name='Arial', size=10, bold=True)
        ws_summary['B14'].font = Font(name='Arial', size=10, bold=True)

        # Tomar últimos 30 días o todos los disponibles
        recent_data = daily_revenue.tail(30)

        chart_row = 15
        for idx, row_data in recent_data.iterrows():
            ws_summary[f'A{chart_row}'] = str(row_data['Fecha_Solo'])
            ws_summary[f'B{chart_row}'] = row_data['Precio Venta']
            chart_row += 1

        # Crear gráfico de líneas
        line_chart = LineChart()
        line_chart.title = "Facturación Diaria (Últimos 30 días)"
        line_chart.style = 13
        line_chart.y_axis.title = 'Revenue (USD)'
        line_chart.x_axis.title = 'Fecha'
        line_chart.height = 12
        line_chart.width = 20

        data = Reference(ws_summary, min_col=2, min_row=14, max_row=14 + len(recent_data))
        cats = Reference(ws_summary, min_col=1, min_row=15, max_row=14 + len(recent_data))
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(cats)

        ws_summary.add_chart(line_chart, "D12")

        # ═══ SECCIÓN 5: GRÁFICO DE GANANCIA POR PRODUCTO ═══
        # Preparar datos para gráfico
        ws_summary['A48'] = 'GANANCIA POR PRODUCTO'
        ws_summary['A48'].font = Font(name='Arial', size=12, bold=True, color="1F4E78")

        chart_row = 49
        for idx, prod_row in top_products.iterrows():
            ws_summary[f'A{chart_row}'] = prod_row['Producto'][:25]
            ws_summary[f'B{chart_row}'] = prod_row['GANANCIA']
            chart_row += 1

        # Crear gráfico de barras
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Top 5 Productos por Ganancia"
        chart1.y_axis.title = 'Ganancia (USD)'

        data = Reference(ws_summary, min_col=2, min_row=48, max_row=48 + len(top_products))
        cats = Reference(ws_summary, min_col=1, min_row=49, max_row=48 + len(top_products))
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.height = 10
        chart1.width = 15

        ws_summary.add_chart(chart1, "D35")

        # Ajustar anchos de columna
        ws_summary.column_dimensions['A'].width = 22
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 12
        ws_summary.column_dimensions['D'].width = 20
        ws_summary.column_dimensions['E'].width = 15
        ws_summary.column_dimensions['F'].width = 15
        ws_summary.column_dimensions['G'].width = 28
        ws_summary.column_dimensions['H'].width = 12

    else:
        ws_summary['A1'] = '📊 DASHBOARD DE VENTAS'
        ws_summary['A1'].font = Font(name='Arial', size=18, bold=True, color="1F4E78")
        ws_summary['A3'] = 'ℹ️ No hay ventas registradas aún'
        ws_summary['A3'].font = Font(name='Arial', size=12)

    # ═══ HOJA 3: VENTAS POR DÍA ═══
    ws_daily = wb.create_sheet("Ventas Por Día", 2)

    if len(df) > 0 and 'Fecha' in df.columns:
        # Preparar datos diarios (solo ventas confirmadas)
        df_copy = df[df['Estado'] != 'cancelled'].copy()
        df_copy['Fecha_parsed'] = pd.to_datetime(df_copy['Fecha'], format='mixed', errors='coerce')
        df_copy = df_copy[df_copy['Fecha_parsed'].notna()].copy()
        df_copy['Fecha_Solo'] = df_copy['Fecha_parsed'].dt.date

        # Agrupar por día
        daily_summary = df_copy.groupby('Fecha_Solo').agg({
            'Producto': 'count',           # Cantidad de ventas
            'Cant': 'sum',                 # Unidades vendidas
            'Precio Venta': 'sum',         # Revenue
            'Total Costo': 'sum',          # Costos
            'GANANCIA': 'sum',             # Ganancia
            'Margen %': 'mean',            # Margen promedio
            'ASIN': 'nunique',             # Productos únicos
            'MKT': 'nunique'               # Marketplaces
        }).reset_index()

        # Renombrar columnas
        daily_summary.columns = [
            'Fecha', 'Ventas', 'Unidades', 'Revenue',
            'Costos Totales', 'Ganancia', 'Margen %',
            'Productos Únicos', 'MKTs Activos'
        ]

        # Ordenar descendente (más reciente primero)
        daily_summary = daily_summary.sort_values('Fecha', ascending=False)

        # Estilos (igual que hoja Ventas)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")

        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

        # Headers
        headers = list(daily_summary.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws_daily.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thick_border

        # Datos
        for row_idx, row_data in enumerate(daily_summary.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_daily.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='Arial', size=10, bold=True)
                cell.border = thick_border
                cell.alignment = Alignment(vertical='center')

        # Formato de números
        for row_idx in range(2, len(daily_summary) + 2):
            # Revenue, Costos, Ganancia (columnas D, E, F)
            for col in [4, 5, 6]:
                ws_daily.cell(row=row_idx, column=col).number_format = '$#,##0.00'
            # Margen %
            ws_daily.cell(row=row_idx, column=7).number_format = '0.00"%"'

        # Fila de TOTALES
        total_row = len(daily_summary) + 2
        ws_daily[f'A{total_row}'] = 'TOTAL'
        ws_daily[f'A{total_row}'].font = Font(name='Arial', size=10, bold=True)
        ws_daily[f'A{total_row}'].border = thick_border

        ws_daily[f'B{total_row}'] = daily_summary['Ventas'].sum()
        ws_daily[f'B{total_row}'].font = Font(name='Arial', size=10, bold=True)
        ws_daily[f'B{total_row}'].border = thick_border

        ws_daily[f'C{total_row}'] = daily_summary['Unidades'].sum()
        ws_daily[f'C{total_row}'].font = Font(name='Arial', size=10, bold=True)
        ws_daily[f'C{total_row}'].border = thick_border

        ws_daily[f'D{total_row}'] = daily_summary['Revenue'].sum()
        ws_daily[f'E{total_row}'] = daily_summary['Costos Totales'].sum()
        ws_daily[f'F{total_row}'] = daily_summary['Ganancia'].sum()
        ws_daily[f'G{total_row}'] = daily_summary['Margen %'].mean()

        # Formato totales (columnas H, I sin datos)
        ws_daily[f'H{total_row}'].border = thick_border
        ws_daily[f'I{total_row}'].border = thick_border

        # Formato totales
        for col in [4, 5, 6]:
            ws_daily.cell(row=total_row, column=col).number_format = '$#,##0.00'
            ws_daily.cell(row=total_row, column=col).font = Font(name='Arial', size=10, bold=True)
            ws_daily.cell(row=total_row, column=col).border = thick_border
        ws_daily.cell(row=total_row, column=7).number_format = '0.00"%"'
        ws_daily.cell(row=total_row, column=7).font = Font(name='Arial', size=10, bold=True)
        ws_daily.cell(row=total_row, column=7).border = thick_border

        # Filtros automáticos
        last_col = get_column_letter(len(headers))
        ws_daily.auto_filter.ref = f"A1:{last_col}{len(daily_summary) + 1}"

        # Freeze panes
        ws_daily.freeze_panes = 'A2'

        # Anchos de columna
        column_widths = {
            'A': 14,   # Fecha
            'B': 10,   # Ventas
            'C': 11,   # Unidades
            'D': 13,   # Revenue
            'E': 14,   # Costos
            'F': 13,   # Ganancia
            'G': 11,   # Margen %
            'H': 16,   # Productos Únicos
            'I': 14    # MKTs Activos
        }

        for col, width in column_widths.items():
            ws_daily.column_dimensions[col].width = width
    else:
        ws_daily['A1'] = 'No hay datos disponibles'
        ws_daily['A1'].font = Font(name='Arial', size=12)

    # ═══ HOJA 4: VENTAS POR MES ═══
    ws_monthly = wb.create_sheet("Ventas Por Mes", 3)

    if len(df) > 0 and 'Fecha' in df.columns:
        # Preparar datos mensuales (solo ventas confirmadas)
        df_copy = df[df['Estado'] != 'cancelled'].copy()
        df_copy['Fecha_parsed'] = pd.to_datetime(df_copy['Fecha'], format='mixed', errors='coerce')
        df_copy = df_copy[df_copy['Fecha_parsed'].notna()].copy()
        df_copy['Mes'] = df_copy['Fecha_parsed'].dt.to_period('M')

        # Agrupar por mes
        monthly_summary = df_copy.groupby('Mes').agg({
            'Producto': 'count',           # Cantidad de ventas
            'Cant': 'sum',                 # Unidades vendidas
            'Precio Venta': 'sum',         # Revenue
            'Total Costo': 'sum',          # Costos
            'GANANCIA': 'sum',             # Ganancia
            'Margen %': 'mean',            # Margen promedio
            'ASIN': 'nunique',             # Productos únicos
            'MKT': 'nunique'               # Marketplaces
        }).reset_index()

        # Convertir período a string legible
        monthly_summary['Mes'] = monthly_summary['Mes'].astype(str)

        # Renombrar columnas
        monthly_summary.columns = [
            'Mes', 'Ventas', 'Unidades', 'Revenue',
            'Costos Totales', 'Ganancia', 'Margen %',
            'Productos Únicos', 'MKTs Activos'
        ]

        # Ordenar descendente (más reciente primero)
        monthly_summary = monthly_summary.sort_values('Mes', ascending=False)

        # Estilos (igual que hoja Ventas)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")

        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

        # Headers
        headers = list(monthly_summary.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws_monthly.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thick_border

        # Datos
        for row_idx, row_data in enumerate(monthly_summary.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_monthly.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='Arial', size=10, bold=True)
                cell.border = thick_border
                cell.alignment = Alignment(vertical='center')

        # Formato de números
        for row_idx in range(2, len(monthly_summary) + 2):
            # Revenue, Costos, Ganancia (columnas D, E, F)
            for col in [4, 5, 6]:
                ws_monthly.cell(row=row_idx, column=col).number_format = '$#,##0.00'
            # Margen %
            ws_monthly.cell(row=row_idx, column=7).number_format = '0.00"%"'

        # Fila de TOTALES
        total_row = len(monthly_summary) + 2
        ws_monthly[f'A{total_row}'] = 'TOTAL'
        ws_monthly[f'A{total_row}'].font = Font(name='Arial', size=10, bold=True)
        ws_monthly[f'A{total_row}'].border = thick_border

        ws_monthly[f'B{total_row}'] = monthly_summary['Ventas'].sum()
        ws_monthly[f'B{total_row}'].font = Font(name='Arial', size=10, bold=True)
        ws_monthly[f'B{total_row}'].border = thick_border

        ws_monthly[f'C{total_row}'] = monthly_summary['Unidades'].sum()
        ws_monthly[f'C{total_row}'].font = Font(name='Arial', size=10, bold=True)
        ws_monthly[f'C{total_row}'].border = thick_border

        ws_monthly[f'D{total_row}'] = monthly_summary['Revenue'].sum()
        ws_monthly[f'E{total_row}'] = monthly_summary['Costos Totales'].sum()
        ws_monthly[f'F{total_row}'] = monthly_summary['Ganancia'].sum()
        ws_monthly[f'G{total_row}'] = monthly_summary['Margen %'].mean()

        # Formato totales (columnas H, I sin datos)
        ws_monthly[f'H{total_row}'].border = thick_border
        ws_monthly[f'I{total_row}'].border = thick_border

        # Formato totales
        for col in [4, 5, 6]:
            ws_monthly.cell(row=total_row, column=col).number_format = '$#,##0.00'
            ws_monthly.cell(row=total_row, column=col).font = Font(name='Arial', size=10, bold=True)
            ws_monthly.cell(row=total_row, column=col).border = thick_border
        ws_monthly.cell(row=total_row, column=7).number_format = '0.00"%"'
        ws_monthly.cell(row=total_row, column=7).font = Font(name='Arial', size=10, bold=True)
        ws_monthly.cell(row=total_row, column=7).border = thick_border

        # Filtros automáticos
        last_col = get_column_letter(len(headers))
        ws_monthly.auto_filter.ref = f"A1:{last_col}{len(monthly_summary) + 1}"

        # Freeze panes
        ws_monthly.freeze_panes = 'A2'

        # Anchos de columna
        column_widths = {
            'A': 14,   # Mes
            'B': 10,   # Ventas
            'C': 11,   # Unidades
            'D': 13,   # Revenue
            'E': 14,   # Costos
            'F': 13,   # Ganancia
            'G': 11,   # Margen %
            'H': 16,   # Productos Únicos
            'I': 14    # MKTs Activos
        }

        for col, width in column_widths.items():
            ws_monthly.column_dimensions[col].width = width
    else:
        ws_monthly['A1'] = 'No hay datos disponibles'
        ws_monthly['A1'].font = Font(name='Arial', size=12)

    # Guardar
    wb.save(DESKTOP_PATH)

    print(f"✅ Excel creado: {DESKTOP_PATH}")
    print(f"   Total ventas: {len(df)}")
    if len(df) > 0:
        print(f"   Ganancia total: ${df['GANANCIA'].sum():,.2f}")

    return DESKTOP_PATH


def upload_to_dropbox(excel_path):
    """Sube el Excel y DBs a Dropbox con auto-refresh de token"""
    try:
        import sys
        from pathlib import Path

        # Add project root to path
        script_dir = Path(__file__).parent
        sys.path.insert(0, str(script_dir))

        # Import dropbox auth utility (handles auto-refresh)
        from dropbox_auth import get_dropbox_client
        import dropbox

        # Get Dropbox client with auto-refresh
        dbx = get_dropbox_client()
        if not dbx:
            print("⚠️  No se pudo obtener cliente Dropbox - Excel no se subió")
            return False

        # Upload Excel
        with open(excel_path, 'rb') as f:
            file_data = f.read()

        dropbox_path = "/VENTAS_ML_V2.xlsx"
        dbx.files_upload(
            file_data,
            dropbox_path,
            mode=dropbox.files.WriteMode.overwrite
        )
        file_size = len(file_data) / 1024  # KB
        print(f"✅ Excel subido a Dropbox: {dropbox_path} ({file_size:.1f} KB)")

        # Upload sales database
        if Path(DB_PATH).exists():
            with open(DB_PATH, 'rb') as f:
                db_data = f.read()
            dbx.files_upload(
                db_data,
                "/sales_tracking.db",
                mode=dropbox.files.WriteMode.overwrite
            )
            db_size = len(db_data) / 1024  # KB
            print(f"✅ DB de ventas subida: /sales_tracking.db ({db_size:.1f} KB)")

        # Upload listings database
        listings_db = "storage/listings_database.db"
        if Path(listings_db).exists():
            with open(listings_db, 'rb') as f:
                listings_data = f.read()
            dbx.files_upload(
                listings_data,
                "/listings_database.db",
                mode=dropbox.files.WriteMode.overwrite
            )
            listings_size = len(listings_data) / 1024  # KB
            print(f"✅ DB de productos subida: /listings_database.db ({listings_size:.1f} KB)")

        return True

    except ImportError as e:
        print(f"⚠️  Error de importación: {e}")
        print("   Instalá: pip3 install dropbox")
        return False
    except Exception as e:
        print(f"⚠️  Error subiendo a Dropbox: {e}")
        return False


if __name__ == "__main__":
    # Siempre generar Excel (para tener archivo actualizado localmente)
    excel_path = create_professional_excel()

    # Solo subir a Dropbox si hay cambios
    if has_sales_changed():
        print("📊 Hay ventas nuevas - Subiendo a Dropbox...")
        upload_to_dropbox(excel_path)
    else:
        print("ℹ️  No hay cambios - Excel no se sube a Dropbox")
